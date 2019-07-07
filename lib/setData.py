import sys
import openpyxl
import numpy as np
import string
from lib import funcData as ecp
from openpyxl.styles.borders import Border,Side

# ワークブックに項目を反映する
def setIndiceXlsx(wb,indice):
	
	sheet = wb.active
	
	i = 0
	for char in indice:
		i = i + 1
		sheet.cell(row = 1, column = i).value = char
	
	# 編集したワークブックを返す
	return wb


# ワークブックにcsvの内容を反映する
def setCsvData(wb,data):
	
	sheet = wb.active
	
	i = 0
	j = 0
	
	for array in data:
		i = i + 1
		j = 0
		for char in array:
			j = j + 1
			sheet.cell(row = i + 1, column = j).value = char
	
	# 編集したワークブックを返す
	return wb
	

# 枠組みを反映する
def setFlameXlsx(wb,flameXlsx):
	
	sheet = wb.active
	intFinalRow = sheet.max_row
	intFinalColumn = sheet.max_column
	
	i = 0
	
	# セルの幅を反映する
	for j in range(intFinalColumn):
		strPos = ecp.RCtoAbs(i + 1,j + 1)
		strColumn = strPos[0:-1]
		sheet.column_dimensions[strColumn].width = flameXlsx.width[j]
	
	# 1行目のセルの高さを反映する
	sheet.row_dimensions[i + 1].height = flameXlsx.heightIndice
	
	# 1行目のセルの背景色を反映する
	setIndiceBgcolor = openpyxl.styles.PatternFill(patternType = 'solid', fgColor = flameXlsx.bgcolorIndice, bgColor = flameXlsx.bgcolorIndice)
	for j in range(intFinalColumn):
		strPos = ecp.RCtoAbs(i + 1,j + 1)
		sheet[strPos].fill = setIndiceBgcolor
	
	# 2行目以降のセルの高さを反映する
	for i in range(1,intFinalRow):
		sheet.row_dimensions[i + 1].height = flameXlsx.heightData
	
	# 2行目以降のセルの背景色を反映する
	setDataBgcolor = openpyxl.styles.PatternFill(patternType = 'solid', fgColor = flameXlsx.bgcolorData,bgColor = flameXlsx.bgcolorData)
	for i in range(1,intFinalRow):
		for j in range(intFinalColumn):
			strPos = ecp.RCtoAbs(i + 1,j + 1) 
			sheet[strPos].fill = setDataBgcolor
	
	# 枠線の反映させる
	borderStyle = flameXlsx.borderStyle
	borderColor = flameXlsx.borderColor
	border = Border(top = Side(style = borderStyle,color = borderColor),left = Side(style = borderStyle,color = borderColor),right = Side(style = borderStyle,color = borderColor),bottom = Side(style = borderStyle,color = borderColor))
	for i in range(intFinalRow):
		for j in range(intFinalColumn):
			strPos = ecp.RCtoAbs(i + 1,j + 1)
			sheet[strPos].border = border
	
	# 編集したワークブックを返す		
	return wb
